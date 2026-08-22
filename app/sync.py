import os
import time
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from sqlalchemy.orm import Session, joinedload, selectinload
from app.database import SessionLocal, engine, Base
from app.models import Client, Principal, Inquiry, Order, Comment, ActivityLog

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CANDIDATE_PATHS = [
    os.path.join(BASE_DIR, "STATUS 2025-2026 .xlsx"),
    os.path.join(BASE_DIR, "STATUS 2025-2026.xlsx"),
    r"G:\My Drive\Salma Elwakeel\Status\STATUS 2025-2026.xlsx",
    r"G:\My Drive\Salma Elwakeel\Status\STATUS 2025-2026 .xlsx",
    r"c:\Users\yassein ahmed\OneDrive\Desktop\CRM\Final-update-to-the-crm-\STATUS 2025-2026 .xlsx",
    r"c:\Users\yassein ahmed\OneDrive\Desktop\Team Eng\STATUS 2025-2026.xlsx"
]

EXCEL_PATH = next((p for p in CANDIDATE_PATHS if os.path.exists(p)), CANDIDATE_PATHS[0])

EURO_PRINCIPALS = [
    'leser', 'bartec', 'sanco', 'as schneider', 'adams', 'dekomte', 'te.ma', 'fht', 'dungs'
]

EGP_INQUIRY_EXCEL_ROWS = {
    80, 84, 87, 99, 128, 136, 137, 138, 166, 180, 204, 209, 210, 224, 228, 
    237, 248, 254, 255, 266, 269, 271, 273, 279, 284, 296, 303, 305, 311, 319,
    # Explicitly labeled in Excel comments/references
    51, 66, 103, 207, 277, 308
}

def clean_str(val):
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        # For numbers loaded as float strings
        val_str = val_str[:-2]
    return val_str

def clean_float(val):
    if pd.isna(val) or val is None:
        return 0.0
    try:
        # Strip out currency symbols if present in number column
        val_str = str(val).replace("$", "").replace("€", "").replace("USD", "").replace("EUR", "").replace("EGP", "").replace(",", "").strip()
        return float(val_str)
    except ValueError:
        return 0.0

def detect_currency(principal_name, val_raw, ref_text="", comment_text="", default_curr="USD", excel_row=None, sheet_name=""):
    if sheet_name == "Inquires" and excel_row is not None and excel_row in EGP_INQUIRY_EXCEL_ROWS:
        return "EGP"

    val_str = str(val_raw or '')
    combined_text = f"{val_str} {ref_text} {comment_text}".upper()
    
    # 1. Explicit EGP markers in text
    if 'EGP' in combined_text or ' L.E' in combined_text or ' LE ' in combined_text or 'POUND' in combined_text:
        return "EGP"

    # 2. Explicit Dollar markers in text
    if 'USD' in combined_text or '$' in combined_text or 'DOLLAR' in combined_text:
        return "USD"
        
    # 3. Explicit Euro markers in text or value cell string
    if '€' in combined_text or 'EUR' in combined_text or 'EURO' in combined_text:
        return "EUR"
        
    # 4. Deduce from Principal name
    p_lower = str(principal_name or '').strip().lower()
    if p_lower:
        for euro_p in EURO_PRINCIPALS:
            if euro_p in p_lower:
                return "EUR"
            
    return default_curr

def clean_date(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, pd.Timestamp):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    # Try parsing different formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    if val_str.isdigit() and len(val_str) == 5:
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=int(val_str))).strftime("%Y-%m-%d")
        except Exception:
            pass
    return val_str

def get_or_create_client(db: Session, name: str):
    name = name.strip()
    if not name:
        name = "Unknown Client"
    client = db.query(Client).filter(Client.name == name).first()
    if not client:
        client = Client(name=name)
        db.add(client)
        db.commit()
        db.refresh(client)
    return client

def get_or_create_principal(db: Session, name: str):
    name = name.strip()
    if not name:
        name = "Unknown Principal"
    principal = db.query(Principal).filter(Principal.name == name).first()
    if not principal:
        principal = Principal(name=name)
        db.add(principal)
        db.commit()
        db.refresh(principal)
    return principal

def import_from_excel():
    start_time = time.time()
    print("=== [START] Starting Excel Import Sync ===")
    db = SessionLocal()
    try:
        # Reset database tables
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)
        print("Database tables recreated.")

        xls = pd.ExcelFile(EXCEL_PATH)
        sheet_names = xls.sheet_names
        print("Sheets found:", sheet_names)

        # In-memory caches for Client and Principal to avoid duplicate lookups
        clients_cache = {}
        principals_cache = {}

        def get_or_create_client_cached(name: str):
            name = name.strip()
            if not name:
                name = "Unknown Client"
            key = name.lower()
            if key not in clients_cache:
                client = db.query(Client).filter(Client.name.ilike(name)).first()
                if not client:
                    client = Client(name=name)
                    db.add(client)
                    db.flush()  # Generate client.id
                clients_cache[key] = client
            return clients_cache[key]

        def get_or_create_principal_cached(name: str):
            name = name.strip()
            if not name:
                name = "Unknown Principal"
            key = name.lower()
            if key not in principals_cache:
                principal = db.query(Principal).filter(Principal.name.ilike(name)).first()
                if not principal:
                    principal = Principal(name=name)
                    db.add(principal)
                    db.flush()  # Generate principal.id
                principals_cache[key] = principal
            return principals_cache[key]

        # In-memory caches for looking up Inquiry by reference
        inquiries_by_ref = {}  # (client_name, principal_name, inquiry_ref) -> Inquiry
        inquiries_by_quot = {} # (client_name, principal_name, quot_ref) -> Inquiry

        # 1. Parse Inquiries Sheets
        inquiries_sheets = {
            "Inquires": "Active",
            "Declined Inquiries": "Declined",
            "Lost Inquiries": "Lost"
        }

        for sheet, status in inquiries_sheets.items():
            if sheet not in sheet_names:
                print(f"Warning: Sheet {sheet} not found in Excel.")
                continue
            
            df = pd.read_excel(xls, sheet_name=sheet, header=7, usecols="A:P")
            # Remove entirely empty rows
            df = df.dropna(how='all')
            print(f"Processing sheet {sheet}: {len(df)} rows found.")

            for idx, row in df.iterrows():
                excel_row = idx + 9
                principal_name = clean_str(row.get("Principal", ""))
                client_name = clean_str(row.get("Client", ""))
                if not principal_name and not client_name:
                    continue  # skip empty/header padding rows

                client = get_or_create_client_cached(client_name)
                principal = get_or_create_principal_cached(principal_name)

                inq_ref = clean_str(row.get("Inquiry Reference"))
                quot_ref = clean_str(row.get("Quotation Reference"))
                comments_text = clean_str(row.get("Comments / Updates", ""))

                raw_val = row.get(" Values") if " Values" in row else row.get("Values")
                val_num = clean_float(raw_val)
                curr = detect_currency(principal_name, raw_val, f"{inq_ref} {quot_ref}", comments_text, excel_row=excel_row, sheet_name=sheet)

                # Determine offer type (Firm vs Budgetary)
                offer_type = "Budgetary" if ("budget" in inq_ref.lower() or "budget" in quot_ref.lower()) else "Firm"

                inq_date_str = clean_date(row.get("Inquiry Date"))
                if inq_date_str:
                    try:
                        inq_year = int(inq_date_str.split('-')[0])
                        if inq_year < 2025:
                            continue  # Skip inquiries prior to 2025
                    except (ValueError, IndexError):
                        pass

                inquiry = Inquiry(
                    inquiry_date=inq_date_str,
                    last_update=clean_date(row.get("Last Update")),
                    due_date=clean_date(row.get(" Due date") if " Due date" in row else row.get("Due date")),
                    principal_id=principal.id,
                    client_id=client.id,
                    inquiry_reference=inq_ref,
                    quotation_reference=quot_ref,
                    value=val_num,
                    currency=curr,
                    offer_type=offer_type,
                    submission_method=clean_str(row.get("Submission Method")),
                    status=status,
                    bid_bond_value=clean_str(row.get("Bid Bond Value")),
                    performance_bond=clean_str(row.get("Performance Bond")),
                    quotation_validity=clean_str(row.get("Quotation Validity")),
                    expiration_date=clean_date(row.get(" Expiration Date") if " Expiration Date" in row else row.get("Expiration Date")),
                    contact_person=clean_str(row.get("Contact Person")),
                    is_deleted=False
                )
                db.add(inquiry)
                db.flush()  # Generate inquiry.id for comments/logs

                # Cache the inquiry for later Won Orders matching
                if status != "Won":
                    c_key = client.name
                    p_key = principal.name
                    if inquiry.inquiry_reference:
                        inquiries_by_ref[(c_key, p_key, inquiry.inquiry_reference)] = inquiry
                    if inquiry.quotation_reference:
                        inquiries_by_quot[(c_key, p_key, inquiry.quotation_reference)] = inquiry

                # Parse and add comments (joined into a single clean paragraph)
                if comments_text:
                    single_para = " ".join([c.strip() for c in comments_text.split('\n') if c.strip()])
                    if single_para:
                        comment = Comment(
                            inquiry_id=inquiry.id,
                            content=single_para,
                            created_at=datetime.utcnow()
                        )
                        db.add(comment)

                # Add initial activity log
                log = ActivityLog(
                    inquiry_id=inquiry.id,
                    action=f"Imported from Excel '{sheet}' sheet with status '{status}'",
                    timestamp=datetime.utcnow()
                )
                db.add(log)

        # 2. Parse Won Order Sheets
        processed_order_inquiry_ids = set()
        orders_by_num = {}   # (client_id, principal_id, order_number_lower) -> (inquiry, order)
        orders_by_ref = {}   # (client_id, principal_id, inq_ref_lower) -> (inquiry, order)
        orders_by_quot = {}  # (client_id, principal_id, quot_ref_lower) -> (inquiry, order)

        order_sheets = ["Orders", "LESER's Orders", " Bartec Orders", " Bartec Orders 2025"]
        for sheet in order_sheets:
            if sheet not in sheet_names:
                print(f"Warning: Sheet {sheet} not found in Excel.")
                continue
            
            df = pd.read_excel(xls, sheet_name=sheet, header=6, usecols="A:W")
            df = df.dropna(how='all')
            print(f"Processing sheet {sheet}: {len(df)} rows.")

            for idx, row in df.iterrows():
                excel_row = idx + 8
                client_name = clean_str(row.get("Client", ""))
                principal_name = clean_str(row.get("Principal", ""))
                if not client_name and not principal_name:
                    continue

                client = get_or_create_client_cached(client_name)
                principal = get_or_create_principal_cached(principal_name)

                inquiry_ref = clean_str(row.get("Client Refrence (Inquiry no.)"))
                quot_ref = clean_str(row.get("Principal Reference (Quotation no.)"))
                inquiry_date = clean_date(row.get("Inquiry Date"))
                comments_text = clean_str(row.get("Comments", ""))
                ord_num = clean_str(row.get("Order Number"))

                tot_val_col = "Total Price" if "Total Price" in row else ("Total Order Value" if "Total Order Value" in row else "Order Value")
                raw_ord_val = row.get(tot_val_col) if tot_val_col in row else row.get("Order Value")
                total_val_num = clean_float(raw_ord_val)
                curr = detect_currency(principal_name, raw_ord_val, f"{inquiry_ref} {quot_ref}", comments_text, excel_row=excel_row, sheet_name=sheet)

                team_comm = clean_str(row.get("TEAM Commisiion", ""))
                pay_status = clean_str(row.get("Payment Status", ""))

                # Check if this order was already processed from an earlier order sheet to prevent duplicates
                ord_num_key = ord_num.lower() if ord_num and ord_num != '-' else ''
                inq_ref_key = inquiry_ref.lower() if inquiry_ref else ''
                quot_ref_key = quot_ref.lower() if quot_ref else ''

                existing_order_match = None
                if ord_num_key:
                    existing_order_match = orders_by_num.get((client.id, principal.id, ord_num_key)) or orders_by_num.get((principal.id, ord_num_key))
                if not existing_order_match and inq_ref_key:
                    existing_order_match = orders_by_ref.get((client.id, principal.id, inq_ref_key)) or orders_by_ref.get((principal.id, inq_ref_key))
                if not existing_order_match and quot_ref_key:
                    existing_order_match = orders_by_quot.get((client.id, principal.id, quot_ref_key)) or orders_by_quot.get((principal.id, quot_ref_key))

                if existing_order_match:
                    # Enrich existing order with any extra details from this sheet
                    ex_inq, ex_ord = existing_order_match
                    if not ex_ord.order_number and ord_num:
                        ex_ord.order_number = ord_num
                    if not ex_ord.order_date and clean_date(row.get("Order Date")):
                        ex_ord.order_date = clean_date(row.get("Order Date"))
                    if (not ex_ord.total_order_value or ex_ord.total_order_value == 0) and total_val_num:
                        ex_ord.total_order_value = total_val_num
                    if not ex_ord.delivery_term and clean_str(row.get("Delivery Term")):
                        ex_ord.delivery_term = clean_str(row.get("Delivery Term"))
                    if not ex_ord.cargo_x and clean_str(row.get("Cargo-X")):
                        ex_ord.cargo_x = clean_str(row.get("Cargo-X"))
                    if not ex_ord.delay_penalty and clean_str(row.get("Delay Penalty")):
                        ex_ord.delay_penalty = clean_str(row.get("Delay Penalty"))
                    if not ex_ord.delivery_period and clean_str(row.get("Delivery Period")):
                        ex_ord.delivery_period = clean_str(row.get("Delivery Period"))
                    if not ex_ord.expected_delivery_date and clean_date(row.get("Expected Delivery Date")):
                        ex_ord.expected_delivery_date = clean_date(row.get("Expected Delivery Date"))
                    if not ex_ord.performance_bond_guarantee and clean_str(row.get("Performance Bond Guarantee")):
                        ex_ord.performance_bond_guarantee = clean_str(row.get("Performance Bond Guarantee"))
                    if not ex_ord.team_commission and team_comm:
                        ex_ord.team_commission = team_comm
                    if not ex_ord.payment_method and clean_str(row.get("Payment method")):
                        ex_ord.payment_method = clean_str(row.get("Payment method"))
                    if not ex_ord.order_confirmation_number and clean_str(row.get("Order Confirmation Number")):
                        ex_ord.order_confirmation_number = clean_str(row.get("Order Confirmation Number"))
                    if not ex_ord.order_confirmations and clean_str(row.get("Order Confirmations.")):
                        ex_ord.order_confirmations = clean_str(row.get("Order Confirmations."))
                    continue

                # Try to find a matching Inquiry in cache
                inquiry = None
                c_name_clean = client.name
                p_name_clean = principal.name
                if inquiry_ref:
                    inquiry = inquiries_by_ref.get((c_name_clean, p_name_clean, inquiry_ref))
                if not inquiry and quot_ref:
                    inquiry = inquiries_by_quot.get((c_name_clean, p_name_clean, quot_ref))

                if inquiry and inquiry.id in processed_order_inquiry_ids:
                    inquiry = None
                
                if not inquiry:
                    # Create base inquiry since none matched
                    inquiry = Inquiry(
                        inquiry_date=inquiry_date,
                        last_update=clean_date(datetime.utcnow()),
                        due_date="",
                        principal_id=principal.id,
                        client_id=client.id,
                        inquiry_reference=inquiry_ref,
                        quotation_reference=quot_ref,
                        value=total_val_num,
                        currency=curr,
                        offer_type="Firm",
                        submission_method="Excel Import",
                        status="Order",
                        is_deleted=False
                    )
                    db.add(inquiry)
                    db.flush()
                else:
                    # Update existing inquiry status to Order
                    inquiry.status = "Order"

                processed_order_inquiry_ids.add(inquiry.id)

                # Deduce Order Status from Payment Status
                if "Supplied" in pay_status or "Paid" in pay_status:
                    ord_status = "Under Payment" if "Paid" in pay_status else "Shipped"
                else:
                    ord_status = "Under Production"

                order = Order(
                    id=inquiry.id,
                    order_number=ord_num,
                    order_date=clean_date(row.get("Order Date")),
                    order_value=clean_float(row.get("Order Value")),
                    additionals=clean_float(row.get("Additionals")),
                    total_order_value=total_val_num,
                    currency=curr,
                    order_confirmation_number=clean_str(row.get("Order Confirmation Number")),
                    team_commission=team_comm,
                    order_confirmations=clean_str(row.get("Order Confirmations.")),
                    delivery_term=clean_str(row.get("Delivery Term")),
                    cargo_x=clean_str(row.get("Cargo-X")),
                    delay_penalty=clean_str(row.get("Delay Penalty")),
                    delivery_period=clean_str(row.get("Delivery Period")),
                    expected_delivery_date=clean_date(row.get("Expected Delivery Date")),
                    performance_bond_guarantee=clean_str(row.get("Performance Bond Guarantee")),
                    payment_method=clean_str(row.get("Payment method")),
                    payment_status=pay_status,
                    order_status=ord_status,
                    source_sheet=sheet
                )
                db.add(order)

                # Register in order lookup maps
                if ord_num_key:
                    orders_by_num[(client.id, principal.id, ord_num_key)] = (inquiry, order)
                    orders_by_num[(principal.id, ord_num_key)] = (inquiry, order)
                if inq_ref_key:
                    orders_by_ref[(client.id, principal.id, inq_ref_key)] = (inquiry, order)
                    orders_by_ref[(principal.id, inq_ref_key)] = (inquiry, order)
                if quot_ref_key:
                    orders_by_quot[(client.id, principal.id, quot_ref_key)] = (inquiry, order)
                    orders_by_quot[(principal.id, quot_ref_key)] = (inquiry, order)

                comments_text = clean_str(row.get("Comments", ""))
                if comments_text:
                    single_para = " ".join([c.strip() for c in comments_text.split('\n') if c.strip()])
                    if single_para:
                        comment = Comment(
                            inquiry_id=inquiry.id,
                            content=single_para,
                            created_at=datetime.utcnow()
                        )
                        db.add(comment)

                # Add log
                log = ActivityLog(
                    inquiry_id=inquiry.id,
                    action=f"Imported Order details from Excel '{sheet}' sheet",
                    timestamp=datetime.utcnow()
                )
                db.add(log)
        
        db.commit()
        elapsed = time.time() - start_time
        print(f"=== [SUCCESS] Excel Import Sync completed successfully in {elapsed:.2f} seconds ===")
        return {"status": "success", "elapsed_seconds": round(elapsed, 2)}
    except Exception as e:
        db.rollback()
        elapsed = time.time() - start_time
        print(f"=== [FAILED] Excel Import Sync failed after {elapsed:.2f} seconds. Error: {e} ===")
        raise e
    finally:
        db.close()

def export_to_excel():
    start_time = time.time()
    print("=== [START] Starting Excel Export Sync ===")
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        # 1. Export Inquiries Sheets
        inquiries_sheets = {
            "Inquires": "Active",
            "Declined Inquiries": "Declined",
            "Lost Inquiries": "Lost"
        }

        for sheet_name, status in inquiries_sheets.items():
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            
            # Unmerge data zone merged cells to avoid MergedCell read-only errors
            max_r = ws.max_row
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= 9:
                    ws.unmerge_cells(range_string=str(mr))

            records = db.query(Inquiry).options(
                joinedload(Inquiry.client),
                joinedload(Inquiry.principal),
                selectinload(Inquiry.comments)
            ).filter(
                Inquiry.status == status,
                Inquiry.is_deleted == False
            ).all()

            for r_idx, req in enumerate(records, 9):
                comments_text = "\n".join([c.content for c in req.comments])
                
                ws.cell(row=r_idx, column=1, value=req.inquiry_date)
                ws.cell(row=r_idx, column=2, value=req.last_update)
                ws.cell(row=r_idx, column=3, value=req.due_date)
                ws.cell(row=r_idx, column=4, value=req.principal.name if req.principal else "")
                ws.cell(row=r_idx, column=5, value=req.client.name if req.client else "")
                ws.cell(row=r_idx, column=6, value=req.inquiry_reference)
                ws.cell(row=r_idx, column=7, value=req.quotation_reference)
                ws.cell(row=r_idx, column=8, value=req.value)
                ws.cell(row=r_idx, column=9, value=req.submission_method)
                ws.cell(row=r_idx, column=10, value=req.status if status == "Active" else status)
                ws.cell(row=r_idx, column=11, value=req.bid_bond_value)
                ws.cell(row=r_idx, column=12, value=req.performance_bond)
                ws.cell(row=r_idx, column=13, value=req.quotation_validity)
                ws.cell(row=r_idx, column=14, value=req.expiration_date)
                ws.cell(row=r_idx, column=15, value=req.contact_person)
                ws.cell(row=r_idx, column=16, value=comments_text)

            # Clear trailing unused old rows if new count is smaller than old max_r
            new_end_row = 9 + len(records)
            if max_r >= new_end_row:
                for row in ws.iter_rows(min_row=new_end_row, max_row=max_r, min_col=1, max_col=16):
                    for cell in row:
                        if type(cell).__name__ != 'MergedCell':
                            cell.value = None

        # 2. Export Order Sheets
        order_sheets = ["Orders", "LESER's Orders", " Bartec Orders", " Bartec Orders 2025"]
        for sheet_name in order_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            
            max_r = ws.max_row
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= 8:
                    ws.unmerge_cells(range_string=str(mr))

            orders = db.query(Order).join(Inquiry).options(
                joinedload(Order.inquiry).joinedload(Inquiry.client),
                joinedload(Order.inquiry).joinedload(Inquiry.principal),
                joinedload(Order.inquiry).selectinload(Inquiry.comments)
            ).filter(
                Inquiry.status.in_(["Won", "Order"]),
                Inquiry.is_deleted == False,
                Order.source_sheet == sheet_name
            ).all()

            for r_idx, o in enumerate(orders, 8):
                comments_text = "\n".join([c.content for c in o.inquiry.comments])
                
                ws.cell(row=r_idx, column=1, value=o.inquiry.client.name if o.inquiry.client else "")
                ws.cell(row=r_idx, column=2, value=o.inquiry.principal.name if o.inquiry.principal else "")
                ws.cell(row=r_idx, column=3, value=o.inquiry.inquiry_date)
                ws.cell(row=r_idx, column=4, value=o.inquiry.inquiry_reference)
                ws.cell(row=r_idx, column=5, value=o.inquiry.quotation_reference)
                ws.cell(row=r_idx, column=6, value=o.order_number)
                ws.cell(row=r_idx, column=7, value=o.order_date)
                ws.cell(row=r_idx, column=8, value=o.order_value)
                ws.cell(row=r_idx, column=9, value=o.additionals)
                
                if sheet_name == "LESER's Orders":
                    ws.cell(row=r_idx, column=10, value=o.total_order_value)
                    ws.cell(row=r_idx, column=11, value=o.order_confirmation_number)
                    ws.cell(row=r_idx, column=12, value=o.order_confirmations)
                    ws.cell(row=r_idx, column=13, value=o.delivery_term)
                    ws.cell(row=r_idx, column=14, value=o.cargo_x)
                    ws.cell(row=r_idx, column=15, value=o.delay_penalty)
                    ws.cell(row=r_idx, column=16, value=o.delivery_period)
                    ws.cell(row=r_idx, column=17, value=o.expected_delivery_date)
                    ws.cell(row=r_idx, column=18, value="")
                    ws.cell(row=r_idx, column=19, value=o.performance_bond_guarantee)
                    ws.cell(row=r_idx, column=20, value=o.payment_method)
                    ws.cell(row=r_idx, column=21, value=o.payment_status)
                    ws.cell(row=r_idx, column=22, value=comments_text)
                else:
                    ws.cell(row=r_idx, column=10, value=o.total_order_value)
                    ws.cell(row=r_idx, column=11, value=o.order_confirmation_number)
                    ws.cell(row=r_idx, column=12, value=o.team_commission)
                    ws.cell(row=r_idx, column=13, value=o.order_confirmations)
                    ws.cell(row=r_idx, column=14, value=o.delivery_term)
                    ws.cell(row=r_idx, column=15, value=o.cargo_x)
                    ws.cell(row=r_idx, column=16, value=o.delay_penalty)
                    ws.cell(row=r_idx, column=17, value=o.delivery_period)
                    ws.cell(row=r_idx, column=18, value=o.expected_delivery_date)
                    ws.cell(row=r_idx, column=19, value="")
                    ws.cell(row=r_idx, column=20, value=o.performance_bond_guarantee)
                    ws.cell(row=r_idx, column=21, value=o.payment_method)
                    ws.cell(row=r_idx, column=22, value=o.payment_status)
                    ws.cell(row=r_idx, column=23, value=comments_text)

            new_end_row = 8 + len(orders)
            max_col_cnt = 22 if sheet_name == "LESER's Orders" else 23
            if max_r >= new_end_row:
                for row in ws.iter_rows(min_row=new_end_row, max_row=max_r, min_col=1, max_col=max_col_cnt):
                    for cell in row:
                        if type(cell).__name__ != 'MergedCell':
                            cell.value = None

        wb.save(EXCEL_PATH)
        elapsed = time.time() - start_time
        print(f"=== [SUCCESS] Excel Export Sync completed successfully in {elapsed:.2f} seconds ===")
        return {"status": "success", "elapsed_seconds": round(elapsed, 2)}
    except Exception as e:
        elapsed = time.time() - start_time
        print(f"=== [FAILED] Excel Export Sync failed after {elapsed:.2f} seconds. Error: {e} ===")
        raise e
    finally:
        db.close()

if __name__ == "__main__":
    # Test importing
    import_from_excel()
